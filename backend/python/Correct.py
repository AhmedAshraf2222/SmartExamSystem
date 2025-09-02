import cv2
import numpy as np
import pandas as pd
import os
import argparse
from pdf2image import convert_from_path
import tempfile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
from scipy.stats import norm

# ========== Part 1: Extract All Correct Answers from Excel ==========
def extract_all_correct_answers(file_path):
    try:
        df = pd.read_excel(file_path, sheet_name='Exam Details', header=None)
        exam_models = {}

        for model_start_col in range(0, 16, 3):
            model_name = df.iloc[2, model_start_col]
            if pd.isna(model_name):
                continue

            questions = []
            answers = []
            for i in range(3, len(df)):
                question = df.iloc[i, model_start_col]
                answer = df.iloc[i, model_start_col + 1]

                if pd.isna(question) or question == 'Questions' or pd.isna(answer) or answer == 'Correct Answer':
                    continue

                questions.append(question)
                try:
                    if isinstance(answer, (int, float)):
                        answers.append(chr(64 + int(answer)))
                    else:
                        answers.append(str(answer).upper())
                except:
                    answers.append(str(answer).upper())

            exam_models[model_name] = dict(zip(questions, answers))

        return exam_models
    except Exception as e:
        print(f"Error extracting correct answers: {e}")
        return None

# ========== Seat and Model Extraction Logic ==========
seat_cols = [1566, 1630, 1693, 1757, 1820, 1884, 1947, 2011, 2074, 2138]
seat_rows = [510, 435, 360, 285] 
seat_centers = [(x, y) for y in seat_rows for x in seat_cols]

def extract_seat_number(thresh, centers):
    digits = []
    for i in range(4):
        group = centers[i*10:(i+1)*10]
        max_ratio = 0
        selected_digit = '_'
        for j, (x, y) in enumerate(group):
            r = 20
            if y - r < 0 or y + r >= thresh.shape[0] or x - r < 0 or x + r >= thresh.shape[1]:
                continue
            roi = thresh[y - r:y + r, x - r:x + r]
            if roi.size == 0:
                continue
            black_pixels = np.sum(roi == 255)
            ratio = black_pixels / roi.size
            if ratio > max_ratio and ratio > 0.3:
                max_ratio = ratio
                selected_digit = str(j)
        digits.append(selected_digit)
    return ''.join(digits) if any(d != '_' for d in digits) else '____'

model_centers = [(1570 + i*63, 655) for i in range(10)]

def extract_model_number(thresh, centers):
    max_ratio = 0
    selected_index = '_'
    for i, (x, y) in enumerate(centers):
        r = 20
        if y - r < 0 or y + r >= thresh.shape[0] or x - r < 0 or x + r >= thresh.shape[1]:
            continue
        roi = thresh[y - r:y + r, x - r:x + r]
        if roi.size == 0:
            continue
        black_pixels = np.sum(roi == 255)
        ratio = black_pixels / roi.size
        if ratio > max_ratio and ratio > 0.3:
            max_ratio = ratio
            selected_index = i
    return str(selected_index + 1) if selected_index != '_' else '_'

# ========== Extract Answers from Sheet ==========
def extract_bubble_sheet(image_path):
    img = cv2.imread(image_path)
    if img is None:
        print(f"Error: Could not load image {image_path}")
        return None
    
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray = cv2.equalizeHist(gray)
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    _, thresh = cv2.threshold(blurred, 80, 255, cv2.THRESH_BINARY_INV)

    results = {}
    results["seat_num"] = extract_seat_number(thresh, seat_centers)
    results["model_no"] = extract_model_number(thresh, model_centers)

    start_xs = [275, 423, 571, 719, 1389, 1537, 1685, 1833]
    start_y = 1068
    step_y = 90
    num_rows = 24
    choices = ['A', 'B', 'C', 'D']

    centers = []
    for col in range(0, len(start_xs), 4):
        for row in range(num_rows):
            y = start_y + row * step_y
            for i in range(4):
                x = start_xs[col + i]
                centers.append((x, y))

    results["answers"] = []
    for i in range(len(centers) // 4):
        q_number = i + 1
        options = centers[i*4:(i+1)*4]
        scores = []
        for j, (x, y) in enumerate(options):
            r = 25
            if y - r < 0 or y + r >= thresh.shape[0] or x - r < 0 or x + r >= thresh.shape[1]:
                scores.append((choices[j], 0, (x, y)))
                continue
            roi = thresh[y - r:y + r, x - r:x + r]
            if roi.shape[0] == 0 or roi.shape[1] == 0:
                scores.append((choices[j], 0, (x, y)))
                continue
            black_pixels = np.sum(roi == 255)
            ratio = black_pixels / roi.size
            scores.append((choices[j], ratio, (x, y)))
        selected = max(scores, key=lambda x: x[1])
        answer = selected[0] if selected[1] > 0.3 else '_'
        results["answers"].append((q_number, answer))

    return results

# ========== Generate Output ==========
def generate_final_output(bubble_results, correct_answers):
    model_no = bubble_results["model_no"]
    model_key = f"Model {model_no}"
    if model_key not in correct_answers:
        print(f"Error: Model {model_no} not found in correct answers")
        return None

    correct_answers_dict = correct_answers[model_key]
    valid_questions = set(correct_answers_dict.keys())
    filtered_answers = [(q, a) for q, a in bubble_results["answers"] if f"Question {q}" in valid_questions]
    student_answers_str = ' '.join([a for q, a in filtered_answers])

    correct_answers_list = [correct_answers_dict.get(f"Question {q}", '_') for q, a in filtered_answers]
    correct_answers_str = ' '.join(correct_answers_list)

    total_questions = len(filtered_answers)
    correct_count = sum(1 for q, a in filtered_answers if correct_answers_dict.get(f"Question {q}", '_').upper() == a.upper())

    score = (correct_count / total_questions) * 100 if total_questions > 0 else 0

    return {
        "Seat Number": bubble_results["seat_num"],
        "Grade": correct_count,
        "Score (%)": f"{round(score, 2)}%",
        "Model": model_no,
        "Total": total_questions,
        "Marked Answers": student_answers_str,
        "Correct Answers": correct_answers_str
    }

# ========== Process PDF File ==========
def process_pdf_file(pdf_path, correct_answers):
    try:
        pages = convert_from_path(pdf_path, dpi=300)
        all_results = []
        print(f"Found {len(pages)} pages")
        for idx, page in enumerate(pages):
            print(f"Processing page {idx + 1}/{len(pages)}...")
            with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp:
                image_path = tmp.name
                page.save(image_path, "JPEG")
            try:
                bubble_results = extract_bubble_sheet(image_path)
                if bubble_results:
                    final_output = generate_final_output(bubble_results, correct_answers)
                    if final_output:
                        all_results.append(final_output)
            finally:
                if os.path.exists(image_path):
                    os.remove(image_path)
            print(f"Done page {idx + 1}")
        return all_results
    except Exception as e:
        print(f"Error processing PDF file: {e}")
        return []

# ========== Apply Styling to Worksheet ==========
def apply_worksheet_styling(ws, include_score=True):
    table_fill = PatternFill(start_color="ECEFF1", end_color="ECEFF1", fill_type="solid")
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    header_font = Font(name="Arial", size=12, color="FFFFFF", bold=True)
    cell_font = Font(name="Arial", size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Header Row Styling
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Data Rows Styling
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            col_letter = cell.column_letter
            header = ws[f"{col_letter}1"].value

            if col_letter == "A":
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            elif header == "Total":
                cell.fill = table_fill
            else:
                cell.fill = table_fill

            cell.font = cell_font
            if header in ['Marked Answers', 'Correct Answers']:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            cell.border = thin_border

    # Auto Column Width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        adjusted_width = (max_length + 2) * 1.1
        ws.column_dimensions[column].width = min(adjusted_width, 80)

    # Score coloring
    if include_score:
        for row in ws.iter_rows(min_row=2):
            score_cell = row[3]  # "Score (%)"
            score = score_cell.value
            if score is not None:
                try:
                    numeric_score = float(str(score).replace('%', ''))
                except:
                    continue
                if numeric_score >= 85:
                    score_cell.fill = PatternFill(start_color="90CAF9", end_color="90CAF9", fill_type="solid")
                elif numeric_score >= 75:
                    score_cell.fill = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")
                elif numeric_score >= 65:
                    score_cell.fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
                else:
                    score_cell.fill = PatternFill(start_color="FFAB91", end_color="FFAB91", fill_type="solid")

# ========== Create Analysis Sheet ==========
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
import matplotlib.pyplot as plt
from scipy.stats import norm
import numpy as np
import pandas as pd
import os
from openpyxl.utils.dataframe import dataframe_to_rows

def create_analysis_sheet(output_path):
    try:
        # Load the existing Excel file
        wb = load_workbook(output_path)
        results_df = pd.read_excel(output_path, sheet_name="Results")

        if "Analysis" in wb.sheetnames:
            wb.remove(wb["Analysis"])
        ws = wb.create_sheet("Analysis")

        ws['A1'] = "Student Grades Analysis Dashboard"
        ws['A1'].font = Font(size=24, bold=True, color="2E86C1")

        temp_files = []
        row_left = 4
        row_right = 4
        step = 30

        # 1. Bar Chart for Student Grades
        plt.figure(figsize=(6, 4))
        bars = plt.bar(results_df["Seat Number"].astype(str),
                       pd.to_numeric(results_df["Score (%)"].str.replace('%', '')),
                       color='#3498DB')
        plt.title('Student Grades Distribution', fontsize=12)
        plt.xlabel('Seat Number', fontsize=10)
        plt.ylabel('Score (%)', fontsize=10)
        plt.xticks(rotation=45, fontsize=8)
        plt.grid(axis='y', linestyle='--', alpha=0.7)

        bar_chart_path = "bar_chart.png"
        plt.savefig(bar_chart_path, bbox_inches='tight', dpi=300)
        plt.close()
        temp_files.append(bar_chart_path)

        ws[f'A{row_left - 1}'] = "1. Student Grades Bar Chart"
        img = Image(bar_chart_path)
        img.width, img.height = 700, 500
        ws.add_image(img, f'A{row_left}')
        row_left += step

        # 2. Histogram of Grades
        plt.figure(figsize=(6, 4))
        scores = pd.to_numeric(results_df["Score (%)"].str.replace('%', ''))
        plt.hist(scores, bins=10, color='#2ECC71', edgecolor='black')
        plt.title('Histogram of Grades', fontsize=12)
        plt.xlabel('Score (%)', fontsize=10)
        plt.ylabel('Number of Students', fontsize=10)
        plt.grid(axis='y', linestyle='--', alpha=0.7)

        hist_path = "histogram.png"
        plt.savefig(hist_path, bbox_inches='tight', dpi=300)
        plt.close()
        temp_files.append(hist_path)

        ws[f'I{row_right - 1}'] = "2. Grades Histogram"
        img = Image(hist_path)
        img.width, img.height = 700, 500
        ws.add_image(img, f'I{row_right}')
        row_right += step

        # 3. Box Plot
        plt.figure(figsize=(5, 3))
        plt.boxplot(scores, vert=False, patch_artist=True,
                    boxprops=dict(facecolor='#F39C12'))
        plt.title('Box Plot of Grades', fontsize=12)
        plt.xlabel('Score (%)', fontsize=10)
        plt.grid(axis='x', linestyle='--', alpha=0.7)

        boxplot_path = "boxplot.png"
        plt.savefig(boxplot_path, bbox_inches='tight', dpi=300)
        plt.close()
        temp_files.append(boxplot_path)

        ws[f'A{row_left - 1}'] = "3. Grades Box Plot"
        img = Image(boxplot_path)
        img.width, img.height = 500, 500
        ws.add_image(img, f'A{row_left}')
        row_left += step

        # 4. Pie Chart
        plt.figure(figsize=(5, 5))
        model_counts = results_df["Model"].value_counts()
        plt.pie(model_counts, labels=model_counts.index, autopct='%1.1f%%',
                colors=['#E74C3C', '#9B59B6', '#3498DB', '#F1C40F'])
        plt.title('Distribution of Students by Model', fontsize=12)

        pie_path = "pie_chart.png"
        plt.savefig(pie_path, bbox_inches='tight', dpi=300)
        plt.close()
        temp_files.append(pie_path)

        ws[f'I{row_right - 1}'] = "4. Models Distribution Pie Chart"
        img = Image(pie_path)
        img.width, img.height = 500, 500
        ws.add_image(img, f'I{row_right}')
        row_right += step

        # 5. Average Score by Model
        plt.figure(figsize=(6, 4))
        avg_scores = results_df.groupby("Model")["Score (%)"]\
            .apply(lambda x: pd.to_numeric(x.str.replace('%', '')).mean())
        avg_scores.plot(kind='bar', color='#16A085')
        plt.title('Average Score by Model', fontsize=12)
        plt.xlabel('Model', fontsize=10)
        plt.ylabel('Average Score (%)', fontsize=10)
        plt.xticks(rotation=0, fontsize=8)
        plt.grid(axis='y', linestyle='--', alpha=0.7)

        avg_model_path = "avg_model.png"
        plt.savefig(avg_model_path, bbox_inches='tight', dpi=300)
        plt.close()
        temp_files.append(avg_model_path)

        ws[f'A{row_left - 1}'] = "5. Average Score by Model"
        img = Image(avg_model_path)
        img.width, img.height = 700, 500
        ws.add_image(img, f'A{row_left}')
        row_left += step

        # 6. Pass Rate by Model
        plt.figure(figsize=(6, 4))
        pass_rates = results_df.groupby("Model")["Score (%)"]\
            .apply(lambda x: (pd.to_numeric(x.str.replace('%', '')) >= 50).mean() * 100)
        pass_rates.plot(kind='bar', color='#D35400')
        plt.title('Pass Rate by Model (%)', fontsize=12)
        plt.xlabel('Model', fontsize=10)
        plt.ylabel('Pass Rate (%)', fontsize=10)
        plt.xticks(rotation=0, fontsize=8)
        plt.grid(axis='y', linestyle='--', alpha=0.7)

        pass_rate_path = "pass_rate.png"
        plt.savefig(pass_rate_path, bbox_inches='tight', dpi=300)
        plt.close()
        temp_files.append(pass_rate_path)

        ws[f'I{row_right - 1}'] = "6. Pass Rate by Model"
        img = Image(pass_rate_path)
        img.width, img.height = 700, 500
        ws.add_image(img, f'I{row_right}')
        row_right += step

        # 7. Gaussian Distribution
        plt.figure(figsize=(6, 4))
        mu, std = norm.fit(scores)
        x = np.linspace(0, 100, 100)

        if std == 0:
            p = np.zeros_like(x)
        else:
            p = norm.pdf(x, mu, std)
        plt.plot(x, p, linewidth=2, color='#8E44AD')
        plt.hist(scores, density=True, alpha=0.6, color='#7FB3D5')
        plt.title(r'Normal Distribution of Grades$\mu=%.1f$, $\sigma=%.1f$' % (mu, std), fontsize=12)
        plt.xlabel('Score (%)', fontsize=10)
        plt.ylabel('Density', fontsize=10)
        plt.grid(linestyle='--', alpha=0.7)

        gaussian_path = "gaussian.png"
        plt.savefig(gaussian_path, bbox_inches='tight', dpi=300)
        plt.close()
        temp_files.append(gaussian_path)

        ws[f'A{row_left - 1}'] = "7. Normal Distribution of Grades"
        img = Image(gaussian_path)
        img.width, img.height = 700, 500
        ws.add_image(img, f'A{row_left}')
        row_left += step

        # 8. Scatter Plot: Seat Number vs Score
        plt.figure(figsize=(6, 4))
        plt.scatter(results_df["Seat Number"], scores, color='#C0392B')
        plt.title('Seat Number vs Score', fontsize=12)
        plt.xlabel('Seat Number', fontsize=10)
        plt.ylabel('Score (%)', fontsize=10)
        plt.grid(linestyle='--', alpha=0.7)

        scatter_path = "scatter.png"
        plt.savefig(scatter_path, bbox_inches='tight', dpi=300)
        plt.close()
        temp_files.append(scatter_path)

        ws[f'I{row_right - 1}'] = "8. Seat Number vs Score"
        img = Image(scatter_path)
        img.width, img.height = 700, 500
        ws.add_image(img, f'I{row_right}')
        row_right += step

        # Summary Stats
        summary_row = max(row_left, row_right) + 1
        ws[f'A{summary_row}'] = "Summary Statistics"
        ws[f'A{summary_row}'].font = Font(size=14, bold=True, color="2E86C1")

        stats_df = pd.DataFrame({
            'Metric': ['Total Students', 'Average Score', 'Minimum Score',
                       'Maximum Score', 'Standard Deviation', 'Pass Rate'],
            'Value': [
                len(results_df),
                f"{scores.mean():.1f}%",
                f"{scores.min():.1f}%",
                f"{scores.max():.1f}%",
                f"{scores.std():.1f}",
                f"{(scores >= 50).mean() * 100:.1f}%"
            ]
        })

        for r in dataframe_to_rows(stats_df, index=False, header=True):
            ws.append(r)

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['I'].width = 25

        wb.save(output_path)
        # تنسيق الخط داخل جدول Summary Statistics
        for row in ws.iter_rows(min_row=summary_row + 1, max_row=summary_row + 6, min_col=1, max_col=2):
            for cell in row:
                cell.font = Font(name='Arial', size=24) 


    finally:
        for img_path in temp_files:
            if os.path.exists(img_path):
                os.remove(img_path)

# ========== Entry Point ==========
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Correct bubble sheets")
    parser.add_argument("--input", required=True, help="Input path (image, folder, or PDF)")
    parser.add_argument("--excel", required=True, help="Path to Excel file with correct answers")
    parser.add_argument("--output", required=True, help="Path to save the output Excel file")
    args = parser.parse_args()

    input_path = args.input
    correct_answers_path = args.excel
    output_path = args.output

    print("Extracting correct answers from Excel...")
    all_correct_answers = extract_all_correct_answers(correct_answers_path)
    if not all_correct_answers:
        print("Error processing correct answers file.")
        exit(1)

    all_outputs = []

    if input_path.lower().endswith(".pdf"):
        print("Processing PDF file...")
        all_outputs = process_pdf_file(input_path, all_correct_answers)

    elif os.path.isdir(input_path):
        print("Processing all images in folder...")
        for filename in os.listdir(input_path):
            if filename.lower().endswith((".jpg", ".jpeg", ".png")):
                image_path = os.path.join(input_path, filename)
                bubble_results = extract_bubble_sheet(image_path)
                if bubble_results:
                    final_output = generate_final_output(bubble_results, all_correct_answers)
                    if final_output:
                        all_outputs.append(final_output)
    else:
        print("Processing image file...")
        bubble_results = extract_bubble_sheet(input_path)
        if bubble_results:
            final_output = generate_final_output(bubble_results, all_correct_answers)
            if final_output:
                all_outputs.append(final_output)

    if all_outputs:
        output_df = pd.DataFrame(all_outputs)
        output_df = output_df.sort_values(by="Seat Number", ascending=True)

        # Create workbook with THREE sheets
        wb = Workbook()
        
        # Sheet 1: Results (With Score %)
        ws1 = wb.active
        ws1.title = "Results"
        sheet1_df = output_df[["Seat Number", "Model", "Grade", "Score (%)", "Total"]]
        for r in dataframe_to_rows(sheet1_df, index=False, header=True):
            ws1.append(r)
        apply_worksheet_styling(ws1, include_score=True)
        
        # Sheet 2: Details (Full Answers)
        ws2 = wb.create_sheet(title="Details")
        sheet2_df = output_df[["Seat Number", "Model", "Grade", "Total", "Marked Answers", "Correct Answers"]]
        for r in dataframe_to_rows(sheet2_df, index=False, header=True):
            ws2.append(r)
        apply_worksheet_styling(ws2, include_score=False)

        # Save initial workbook first (required for analysis)
        wb.save(output_path)
        
        # Sheet 3: Analysis Dashboard
        print("Creating Analysis Dashboard...")
        try:
            create_analysis_sheet(output_path)
            print("Analysis dashboard created successfully!")
        except Exception as e:
            print(f"Error creating analysis sheet: {e}")

        print(f"\nFinal results saved to: {output_path}")
        print("Sheets created:")
        print("- Results: Summary with scores")
        print("- Details: Full answers data")
        print("- Analysis: Visual dashboard")
    else:
        print("No valid results extracted.")
        exit(1)